import openpyxl
import csv
import sys

def main():
    if len(sys.argv) != 3:
        print("Uso: python Criar Lista de Sistemas.py arquivo_de_entrada_dados.xlsx diretorio_de_saida")
        return

    arquivo_entrada = sys.argv[1]
    diretorio_saida = sys.argv[2]

    # Abrir o arquivo Excel original
    workbook = openpyxl.load_workbook(arquivo_entrada, data_only=True)
    ws_original = workbook["Lista de Plantas e Sistemas"]

    # Nome do arquivo CSV de saída
    csv_file_path = diretorio_saida + "\\Lista_de_Sistemas.csv"

    # Dicionário para verificar duplicatas
    unique_dict = {}

    # Cabeçalhos para o arquivo CSV
    headers = ["ObjectType", "Name", "City", "Neighborhood", "CommandType", "Company", "CompanyAcronym",
               "ComponentId", "ConditionName", "Contract", "ControlCenterArea", "District", "DocString",
               "Organization", "PathVolume", "Region", "State", "StateAcronym", "WaterType", 
               "PathContainer", "PathName", "AlarmVerify", "ID", "IsAlarmArea"]

    # Abrir o arquivo CSV para escrita
    with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)

        # Escrever os cabeçalhos no arquivo CSV
        writer.writerow(headers)

        # Iterar sobre as linhas na planilha de lista e sistemas
        for i in range(2, ws_original.max_row + 1):
            # Obter os valores das colunas relevantes para a linha atual
            prefixo_sistema = str(ws_original.cell(row=i, column=ws_original["M" + str(i)].column).value)  # Prefixo Sistema = Água ou Esgoto
            valor_path_name = str(ws_original.cell(row=i, column=ws_original["S" + str(i)].column).value)  # PathName do Sistema
            split_path_name = valor_path_name.split(".")
            path_container = ".".join(split_path_name[:-1])  # PathContainer sem a última parte
            nome_prj = str(ws_original.cell(row=i, column=ws_original["AO" + str(i)].column).value)  # Onde está o arquivo onde o sistema deve ser criado
            organization = str(ws_original.cell(row=i, column=4).value)  # Organização
            contract = str(ws_original.cell(row=i, column=6).value)  # Contrato
            region = str(ws_original.cell(row=i, column=9).value)  # Regional
            State = "Rio Grande do Sul"  # Corsan é RS
            StateAcronym = "RS"  # Corsan é RS

            if valor_path_name != "Impossível criar sistema com a coluna pendente" and valor_path_name != "":
                split_path_name = valor_path_name.split(".")
                chave_unica = valor_path_name + nome_prj

                if chave_unica not in unique_dict:
                    unique_dict[chave_unica] = ""

                    # Determinar o valor de "AlarmVerify"
                    alarm_verify = "True" if prefixo_sistema == "Agua" else "True"

                    # Preencher as células para a nova linha do CSV
                    new_row_values = [
                        "WaterDistributionNetwork",  # ObjectType
                        split_path_name[-1],  # Name
                        str(ws_original.cell(row=i, column=10).value),  # Cidade
                        str(ws_original.cell(row=i, column=15).value),  # Neighborhood
                        "",  # CommandType
                        str(ws_original.cell(row=i, column=5).value),  # Company
                        str(ws_original.cell(row=i, column=8).value),  # CompanyAcronym
                        "",  # ComponentId
                        "",  # ConditionName
                        contract,  # Contract
                        "",  # ControlCenterArea
                        "",  # District
                        "",  # DocString
                        organization,  # Organization
                        nome_prj,  # PathVolume
                        region,  # Region
                        State,  # State
                        StateAcronym,  # StateAcronym
                        "1" if prefixo_sistema == "Agua" else "2",  # WaterType
                        path_container,  # PathContainer
                        valor_path_name,  # PathName
                        alarm_verify,  # AlarmVerify
                        "{00000000-0000-0000-0000-000000000000}",  # ID
                        alarm_verify  # IsAlarmArea
                    ]
                    writer.writerow(new_row_values)

    # Informar ao usuário que o processo foi concluído
    print("Documento CSV de sistemas com sucesso!")

if __name__ == "__main__":
    main()
