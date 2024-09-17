import openpyxl
import csv
import sys


def main():
    if len(sys.argv) != 3:
        print("Uso: python Criar Lista de Plantas.py arquivo_de_entrada_dados.xlsx diretorio_de_saida")
        return

    arquivo_entrada = sys.argv[1]
    diretorio_saida = sys.argv[2]

    # Abrir o arquivo Excel original
    workbook = openpyxl.load_workbook(arquivo_entrada, data_only=True)
    ws_original = workbook["Lista de Plantas e Sistemas"]

    # Restante do seu código...
    # Certifique-se de atualizar qualquer referência à planilha original para usar ws_original

    # Nome do arquivo CSV de saída
    csv_file_path = diretorio_saida + "\\Lista_de_Plantas.csv"


    # Dicionário para verificar duplicatas
    unique_dict = {}

    # Cabeçalhos para o arquivo CSV - Adicionando novas colunas
    headers = ["ObjectType", "Name", "City", "CommandName", "CommandType", "Company", "CompanyAcronym",
               "ComponentId", "ConditionName", "Contract", "ControlCenterArea", "District", "ShorName",
               "Organization", "PathVolume", "Region", "State", "StateAcronym", "WaterType", "PathContainer",
               "PathName", "AlarmVerify", "ID", "IsAlarmArea", "Latitude", "Longitude", "Address", "ControlCenterSubarea"]

    # Abrir o arquivo CSV para escrita
    with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)

        # Escrever os cabeçalhos no arquivo CSV
        writer.writerow(headers)

        # Iterar sobre as linhas na planilha original
        for i in range(2, ws_original.max_row + 1):
            # Ignorar linhas com "Coluna Pendente" em valor_path_name (coluna "R")
            valor_path_name = str(ws_original.cell(row=i, column=ws_original["S" + str(i)].column).value)
            if valor_path_name == "Impossível criar sistema com a coluna pendente":
                continue
            
            # Obter os valores das colunas relevantes para a linha atual
            # (Incluindo as novas colunas Latitude, Longitude, Address, ControlCenterSubarea e a correção para DocString)
            caminho_planta = str(ws_original.cell(row=i, column=ws_original["AD"+ str(i)].column).value)
            nome_prj = str(ws_original.cell(row=i, column=ws_original["AO"+ str(i)].column).value)
            organization = str(ws_original.cell(row=i, column=4).value)
            contract = str(ws_original.cell(row=i, column=6).value)
            region = str(ws_original.cell(row=i, column=9).value)
            city = str(ws_original.cell(row=i, column=10).value)
            company = str(ws_original.cell(row=i, column=5).value)
            prefixo_sistema = str(ws_original.cell(row=i, column=ws_original["M"+ str(i)].column).value)
            latitude = str(ws_original.cell(row=i, column=ws_original["Z" + str(i)].column).value)
            longitude = str(ws_original.cell(row=i, column=ws_original["AA" + str(i)].column).value)
            address = str(ws_original.cell(row=i, column=ws_original["AB" + str(i)].column).value)

            # Verifica se latitude, longitude e address estão vazios e define como string vazia se estiverem
            latitude = latitude if latitude and latitude.strip() != "None" else ""
            longitude = longitude if longitude and longitude.strip() != "None" else ""
            address = address if address and address.strip() != "None" else ""
            control_center_subarea = str(ws_original.cell(row=i, column=ws_original["Q"+ str(i)].column).value)
            ShortName = str(ws_original.cell(row=i, column=ws_original["AF"+ str(i)].column).value)
            
            State = "Rio Grande do Sul"
            StateAcronym = "RS"
            
            split_caminho_planta = caminho_planta.split(".")
            chave_unica = caminho_planta + nome_prj

            if chave_unica not in unique_dict:
                unique_dict[chave_unica] = None

                # Preencher as células para a nova linha do CSV
                new_row_values = [
                    "WaterStation",  # ObjectType
                    split_caminho_planta[-1],  # Name
                    city,
                    "",  # CommandName
                    "",  # CommandType
                    company,
                    str(ws_original.cell(row=i, column=8).value),  # CompanyAcronym
                    "",  # ComponentId
                    "",  # ConditionName
                    contract,
                    "",  # ControlCenterArea
                    "",  # District
                    ShortName,  # DocString, corrigido para vir da coluna AE
                    organization,
                    nome_prj,  # PathVolume
                    region,
                    State,
                    StateAcronym,
                    "1" if prefixo_sistema == "Agua" else "2",  # WaterType
                    valor_path_name,  # PathContainer
                    caminho_planta,
                    "True",  # AlarmVerify
                    "{00000000-0000-0000-0000-000000000000}",  # ID
                    "True",  # IsAlarmArea
                    latitude,  # Latitude
                    longitude,  # Longitude
                    address,  # Address
                    control_center_subarea  # ControlCenterSubarea
                ]
                writer.writerow(new_row_values)

    # Informar ao usuário que o processo foi concluído
    print("Documento CSV de plantas criado com sucesso!")


if __name__ == "__main__":
    main()