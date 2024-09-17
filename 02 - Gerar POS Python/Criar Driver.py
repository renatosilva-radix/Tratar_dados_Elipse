import openpyxl
import csv
import re
import sys

def main(arquivo_funcionalidades, nome_planilha):
    arquivo_entrada = arquivo_funcionalidades  # Use o arquivo de funcionalidades fornecido
    nome_planilha = nome_planilha  # Use o nome da planilha fornecido
    workbook = openpyxl.load_workbook(arquivo_entrada, data_only=True)
    ws_original = workbook[nome_planilha]  # Use o nome da planilha fornecido

    # Nome do arquivo CSV de saída
    csv_file_path_medidas = diretorio_saida + "\\Lista_Driver.csv"

    # Cabeçalhos para o arquivo CSV de medidas
    headers_medidas = [
        "Protocolo", "Name", "PathContainer", "IOFolder",
        "PathName", "ParamDevice", "ParamItem",  
        "UseBitFields", "EnableScalling", "EuLow", "EuHigh",
        "DeviceLow", "DeviceHigh", "PathVolume"
    ]

    # Conjunto para rastrear PathName únicos
    unique_path_names = set()

# Abrir o arquivo CSV para escrita das medidas
    with open(csv_file_path_medidas, mode='w', newline='', encoding='utf-8-sig') as file_medidas:

        writer_medidas = csv.writer(file_medidas)
        writer_medidas.writerow(headers_medidas)

        # Iterar sobre as linhas na planilha de equipamentos e funcionalidades
        for i in range(2, ws_original.max_row + 1):
            PathName = str(ws_original.cell(row=i, column=ws_original["AM" + str(i)].column).value)
            Protocolo = str(ws_original.cell(row=i, column=ws_original["AF" + str(i)].column).value)

            # Verifica se o PathName é válido e não é duplicado
            if not PathName or PathName in unique_path_names or PathName in ["#N/D", "#N/A", "None", "Em definição", None]:
                continue
            unique_path_names.add(PathName)


            Name = str(ws_original.cell(row=i, column=33).value)  # Coluna  "AG"
            PathContainer = str(ws_original.cell(row=i, column=34).value)  # Coluna "AH"
            IOFolder = str(ws_original.cell(row=i, column=35).value)  # Coluna "AI"
            ParamItem = str(ws_original.cell(row=i, column=41).value) # Coluna "AO"
            UseBitFields = str(ws_original.cell(row=i, column=49).value)  # Coluna "AW"
            EnableScalling = str(ws_original.cell(row=i, column=50).value)  # Coluna "AX"
            EuLow = str(ws_original.cell(row=i, column=51).value)  # Coluna "AY"
            EuHigh = str(ws_original.cell(row=i, column=52).value)  # Coluna "AZ"
            DeviceLow = str(ws_original.cell(row=i, column=53).value)  # Coluna "BA"
            DeviceHigh = str(ws_original.cell(row=i, column=54).value)  # Coluna "BB"
            PathVolume = str(ws_original.cell(row=i, column=18).value)  # Coluna "R"

            # Dentro do loop, quando você define ParamDevice:
            ParamDevice = str(ws_original.cell(row=i, column=ws_original["AN" + str(i)].column).value)
    
            # Adiciona um caractere tab no início para evitar formatação automática no Excel
            ParamDevice = f"\t{ParamDevice}"

        
                    
            # Escreve a nova linha no arquivo CSV
            new_row_medidas = [
                # Inclui todas as colunas previamente definidas
                Protocolo, Name, PathContainer, IOFolder, PathName, ParamDevice, ParamItem,  
                UseBitFields, EnableScalling, EuLow, EuHigh,
                DeviceLow, DeviceHigh, PathVolume
            ]
            writer_medidas.writerow(new_row_medidas)

    # Informar ao usuário que o processo foi concluído
    print("Documento CSV de driver criado com sucesso!")

if len(sys.argv) != 4:
    print("Uso: python Criar Driver.py arquivo_de_entrada_dados.xlsx nome_planilha diretorio_de_saida")
    sys.exit(1)

arquivo_funcionalidades = sys.argv[1]
nome_planilha = sys.argv[2]
diretorio_saida = sys.argv[3]

main(arquivo_funcionalidades, nome_planilha)