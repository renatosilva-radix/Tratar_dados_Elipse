import openpyxl
import csv
import sys

def main(arquivo_funcionalidades, nome_planilha):
    arquivo_entrada = arquivo_funcionalidades  # Use o arquivo de funcionalidades fornecido
    nome_planilha = nome_planilha  # Use o nome da planilha fornecido
    workbook = openpyxl.load_workbook(arquivo_entrada, data_only=True)
    ws_original = workbook[nome_planilha]  # Use o nome da planilha fornecido

    # Nome do arquivo CSV de saída
    csv_file_path = diretorio_saida + "\\Lista_de_Equipamentos.csv"

    # Dicionário para verificar duplicatas
    unique_dict = {}

    # Cabeçalhos para o arquivo CSV - Adicionando novas colunas
    headers = ["ObjectType", "Name", "Fase", "ShortName","Caption","Type","PathContainer", "PathName", "PathVolume"]

    # Abrir o arquivo CSV para escrita
    with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)

        # Escrever os cabeçalhos no arquivo CSV
        writer.writerow(headers)

        # Iterar sobre as linhas na planilha original
        for i in range(2, ws_original.max_row + 1):
            # Obter os valores das colunas relevantes para a linha atual
            nome_Eqp = str(ws_original.cell(row=i, column=ws_original["I"+ str(i)].column).value)
            Fase = str(ws_original.cell(row=i, column=ws_original["M"+ str(i)].column).value)
            nome_prj = str(ws_original.cell(row=i, column=ws_original["R"+ str(i)].column).value)
            coluna_d = str(ws_original.cell(row=i, column=ws_original["D"+ str(i)].column).value)
            valor_path_name = str(ws_original.cell(row=i, column=ws_original["Q"+ str(i)].column).value)
            
            if valor_path_name == "Coluna Pendente":
                continue
            
            # Definir ObjectType com base na coluna D
            if coluna_d == "Area (WaterGenericItem)":
                ObjectType = "WaterGenericItem"
                Type = "Area"
                Caption = Type
                ShortName = str(ws_original.cell(row=i, column=ws_original["T"+ str(i)].column).value)
            else:
                ObjectType = coluna_d
                Type = str(ws_original.cell(row=i, column=ws_original["U"+ str(i)].column).value)
                Caption = Type
                ShortName = str(ws_original.cell(row=i, column=ws_original["T"+ str(i)].column).value)
            
            # Definir PathContainer removendo a última posição de valor_path_name
            path_parts = valor_path_name.rsplit(".", 1)  # Divide a string pela última ocorrência de "."
            if len(path_parts) > 1:
                PathContainer = path_parts[0] 
            else:
                PathContainer = valor_path_name 

            # Verificar se todos os valores da linha são None
            if all(v is None for v in [ObjectType,nome_Eqp,Fase ,ShortName ,Caption ,Type ,PathContainer ,valor_path_name ,nome_prj]):
                continue  # Ignorar esta linha se todos os valores forem None

            # Substituir valores None por "Informação não encontrada"
            ObjectType = ObjectType if ObjectType is not None else "Informação não encontrada"
            nome_Eqp = nome_Eqp if nome_Eqp is not None else "Informação não encontrada"
            Fase = Fase if Fase is not None else "Informação não encontrada"
            ShortName = ShortName if ShortName is not None else "Informação não encontrada"
            Caption = Caption if Caption is not None else "Informação não encontrada"
            Type = Type if Type is not None else "Informação não encontrada"
            PathContainer = PathContainer if PathContainer is not None else "Informação não encontrada"
            nome_prj = nome_prj if nome_prj is not None else "Informação não encontrada"
            valor_path_name = valor_path_name if valor_path_name is not None else "Informação não encontrada"
            
            split_caminho_Eqp = valor_path_name.split(".")
            chave_unica = valor_path_name + nome_prj

            if chave_unica not in unique_dict:
                unique_dict[chave_unica] = None

                # Preencher as células para a nova linha do CSV
                new_row_values = [
                    ObjectType,  # ObjectType
                    nome_Eqp,  # Name
                    Fase,  # Fase
                    ShortName,  # ShortName
                    Caption,  # ShortName
                    Type,  # ShortName
                    PathContainer,  # PathContainer
                    valor_path_name,  # PathName
                    nome_prj,  # PathVolume
                ]
                writer.writerow(new_row_values)

    # Informar ao usuário que o processo foi concluído
    print("Documento CSV de equipamentos criado com sucesso!")

if len(sys.argv) != 4:
    print("Uso: python Criar Lista de Equipamentos.py arquivo_de_entrada_dados.xlsx nome_planilha diretorio_de_saida")
    sys.exit(1)

arquivo_funcionalidades = sys.argv[1]
nome_planilha = sys.argv[2]
diretorio_saida = sys.argv[3]

main(arquivo_funcionalidades, nome_planilha)
