import openpyxl
import csv
import re
import sys


def main(arquivo_funcionalidades, nome_planilha, diretorio_saida):
    arquivo_entrada = arquivo_funcionalidades  # Use o arquivo de funcionalidades fornecido
    nome_planilha = nome_planilha  # Use o nome da planilha fornecido
    workbook = openpyxl.load_workbook(arquivo_entrada, data_only=True)
    ws_original = workbook['Equipamentos e Funcionalidades']  # Use o nome da planilha fornecido

    # Nome do arquivo CSV de saída
    csv_file_path_medidas = diretorio_saida + "\\Lista_de_Medidas.csv"

    # Cabeçalhos para o arquivo CSV de medidas
    headers_medidas = [
        "ObjectType", "Name", "MeasurementType", "Valor",
        "Unit", "ActiveSource", "StudyInputSource", "PathName", "PathVolume",
        "CaminhoEquipamento", "Sistema", "Conversao", "Tag", "Modelo", "OperateFeedback", "ProcessVariableMeasurement"
    ]

    # Conjunto para rastrear PathName únicos
    unique_path_names = set()

    # Abrir o arquivo CSV para escrita das medidas
    with open(csv_file_path_medidas, mode='w', newline='', encoding='utf-8-sig') as file_medidas:
        writer_medidas = csv.writer(file_medidas, delimiter=';')
        writer_medidas.writerow(headers_medidas)

        # Iterar sobre as linhas na planilha de equipamentos e funcionalidades
        for i in range(2, ws_original.max_row + 1):
            coluna_p_preenchida = str(ws_original.cell(row=i, column=ws_original["P" + str(i)].column).value)
            path_name = str(ws_original.cell(row=i, column=ws_original["S" + str(i)].column).value)
            object_type = str(ws_original.cell(row=i, column=ws_original["O" + str(i)].column).value)
            equipment_type = str(ws_original.cell(row=i, column=4).value)  # Coluna 'D'
            # Verifica se o PathName é válido e não é duplicado
            if not path_name or path_name in unique_path_names or path_name in ["#N/D", "#N/A", "None", None]:
                continue
            unique_path_names.add(path_name)

            # Define Name com base nas colunas P e S
            if coluna_p_preenchida:
                path_parts = path_name.split('.')
                if "AlarmeAnalogico" in object_type and len(path_parts) > 1:
                    # Se o ObjectType contém "AlarmeAnalogico", o Name é a penúltima posição de PathName
                    name = path_parts[-2]
                else:
                    # Se a coluna P está preenchida, o Name é a última posição de PathName
                    name = path_parts[-1]
            else:
                # Se a coluna P não está preenchida, mantém o valor original da coluna K para Name
                name = ws_original.cell(row=i, column=ws_original["K"].column).value

            object_type = str(ws_original.cell(row=i, column=15).value)  # Coluna "O"
            valor = str(ws_original.cell(row=i, column=3).value)  # Coluna para "Valor"
            equip_path = str(ws_original.cell(row=i, column=17).value)  # Coluna "Q"
            path_volume = str(ws_original.cell(row=i, column=18).value)  # Coluna "R"
            Tag_Driver = str(ws_original.cell(row=i, column=39).value)
            UsoBit = str(ws_original.cell(row=i, column=46).value)
            modelo = str(ws_original.cell(row=i, column=5).value) if str(ws_original.cell(row=i, column=5).value) != 'None' else ""  # Coluna "E"
            operateFeedback = str(ws_original.cell(row=i, column=26).value) if str(ws_original.cell(row=i, column=26).value) != 'None' else ""
            processVariableMeasurements = str(ws_original.cell(row=i, column=27).value) if str(ws_original.cell(row=i, column=27).value) != 'None' else ""


            # Define CaminhoEquipamento baseado em ObjectType
            if ".Terminal" in path_name:
                # Se ".Terminal" está presente em path_name, define caminho_equipamento para ".Terminal" seguido do número após "Terminal"
                index_terminal = path_name.index(".Terminal")
                terminal_index = re.search(r'\d+', path_name[index_terminal:]).group()
                caminho_equipamento = f"{equip_path}.Terminal{terminal_index}"
            elif "Command" in object_type:
                caminho_equipamento = f"{equip_path}.Commands"
            else:
                # Se ".Terminal" não está presente em path_name e "Command" não está presente em object_type, define caminho_equipamento como ".Measurements"
                caminho_equipamento = f"{equip_path}.Measurements"

            # Define Sistema a partir de CaminhoEquipamento e Nome da Planta
            caminho_equipamento = ws_original.cell(row=i, column=17).value
            caminho_equipamento_parts = caminho_equipamento.split('.')
            nome_planta = str(ws_original.cell(row=i, column=3).value)  # Coluna "C" na planilha original

            # Inicializa o sistema como vazio
            sistema = ''

            # Percorre cada parte do caminho do equipamento
            for parte in caminho_equipamento_parts:
                # Adiciona ao sistema até encontrar um elemento que começa com "PL_"
                if parte.startswith("PL_"):
                    break
                sistema += parte + '.'

            # Remove o último ponto extra
            if sistema.endswith('.'):
                sistema = sistema[:-1]

            active_source = 1

            # Regras para StudyInputSource
            # Aqui você pode adicionar a lógica para definir StudyInputSource conforme necessário
            # Por padrão, estamos definindo como 0
            study_input_source = 0

            conversao = "(default)"  # Conversão vazia para WaterAnalogMeasurement

            # Assume-se que Tag Driver e UsoBit são da coluna 39 e 46, respectivamente
            Tag_Driver = str(ws_original.cell(row=i, column=39).value)
            UsoBit = str(ws_original.cell(row=i, column=46).value)

            # Regras para a coluna Tag
            if name.startswith("Word_Calculo_"):
                # Encontra o outro registro correspondente
                caminho_equipamento_calculo = caminho_equipamento.rsplit(".", 1)[0] + "." + name.split("_")[-1]
                for j in range(2, ws_original.max_row + 1):
                    if ws_original.cell(row=j, column=17).value == caminho_equipamento_calculo:
                        tag = ws_original.cell(row=j, column=39).value
                        break
                else:
                    tag = Tag_Driver  # Usar o Tag Driver caso nenhum correspondente seja encontrado
            elif name.startswith("FQIT"):
                for j in range(2, ws_original.max_row + 1):
                    if ws_original.cell(row=j, column=17).value == caminho_equipamento:
                        tag = f"(IO:{ws_original.cell(row=j, column=39).value}.Value / 4000) * 360"
                        break
                else:
                    tag = Tag_Driver  # Usar o Tag Driver caso nenhum correspondente seja encontrado
            else:
                if UsoBit.isdigit():  # Verifica se UsoBit é um número e cria uma tag específica com Bit
                    at_number = int(UsoBit)
                    at_formatted = f"0{at_number}" if at_number < 10 else str(at_number)
                    tag = f"{Tag_Driver}.Bit{at_formatted}"
                else:
                    tag = Tag_Driver  # Usar o Tag Driver diretamente se não for um caso de Bit

            if name == "Y" or name == "V_Y":
                if "WaterPump" in equipment_type or "WaterSubmersibleWellPump" in equipment_type:
                    measurement_type = "*PumpState"
                    unit = ''
                elif "WaterShutoffValve" in equipment_type:
                    measurement_type = "*ValveState"
                    unit = ''
                else:
                    measurement_type = name
                    unit = ''
            # Aplicando a lógica de match para definir MeasurementType e Unit
            elif re.match(r'M_LIT_\d+', name) or re.match(r'LIT_\d+', name) or re.match(r'J_LIT_\d+', name) or re.match(
                    r'\bM_LIT\b', name) or re.match(r'\bJ_LIT\b', name):
                measurement_type = "*WaterLevel"
                unit = "Percent"
            elif re.match(r'\bPIT\b', name) or re.match(r'M_PIT\b', name) or re.match(r'J_PIT\b', name):
                measurement_type = "*Pressure"
                unit = "mH2O"
            elif re.match(r'\bFIT\b', name) or re.match(r'M_FIT\b', name) or re.match(r'J_FIT\b', name):
                measurement_type = "*FlowRate"
                unit = "LPS"
            elif re.match(r'\bFQIT\b', name) or re.match(r'M_FQIT\b', name) or re.match(r'J_FQIT\b', name):
                measurement_type = "*WaterVolumePositiveFlow"
                unit = "CubicMeters"
            elif name == "SIT" or name == "SIT_1":
                measurement_type = "SIT"
                unit = "Hz"
            elif name == "SIT_2":
                measurement_type = "*PumpSpeed"
                unit = "rpm"
            elif name.startswith("EIT"):
                measurement_type = name  # Measurement type é o próprio Name
                unit = "V"
            elif name.startswith("IIT"):
                measurement_type = name  # Measurement type é o próprio Name
                unit = "A"
            elif name == "ATV" or name == "JIT_ATV":
                measurement_type = name  # Measurement type é o próprio Name
                unit = "KW"
            elif name == "JIT_RTV":
                measurement_type = name  # Measurement type é o próprio Name
                unit = "KVAR"
            elif name == "JIT_FP" or name == "Y_SW":
                measurement_type = name  # Measurement type é o próprio Name
                unit = "Adimensional"
            elif name == "Y_ON_Q":
                measurement_type = name  # Measurement type é o próprio Name
                unit = "HHMMSS"
            elif name == "CC_CR" or name == "CC_CE":
                measurement_type = name  # Measurement type é o próprio Name
                unit = "Adimensional"
            elif name == "CC_TX":
                measurement_type = name  # Measurement type é o próprio Name
                unit = "Percent"
            elif name == "TIT":
                measurement_type = name  # Measurement type é o próprio Name
                unit = "C"
            elif name == "MGE_SIT":
                measurement_type = name  # Measurement type é o próprio Name
                unit = "Hz"
            elif "MGE" in name and "_COD_Y" in name:
                measurement_type = "MGE_COD_Y"  # Measurement type é sempre MGE_COD_Y
                unit = "Adimensional"
            elif "S" == name:
                measurement_type = "*PumpState"  # Measurement type é sempre MGE_COD_Y
                unit = ''
            elif name== "MGEE_ATV" or name == "MGE_ATV":
                measurement_type="E_ATV"
                unit = ''
            elif name == "M_PSL_ON_SP":
                measurement_type="PSL_ON_SP"
                unit = ''
            elif name == "M_PSL_OFF_SP":
                measurement_type="PSL_OFF_SP"
                unit = ''
            elif name in ["LR", "SIT", "SIT2"]:
                measurement_type = "*PumpState"  # Measurement type é sempre MGE_COD_Y
                unit = ''
            else:
                measurement_type = {
                }.get(name, name)
                unit = {
                    "LIT": "Percent",
                }.get(name, "")

            # Escreve a nova linha no arquivo CSV
            new_row_medidas = [
                # Inclui todas as colunas previamente definidas
                object_type, name, measurement_type, valor,
                unit, active_source, study_input_source, path_name, path_volume,
                caminho_equipamento, sistema, conversao, tag, modelo, operateFeedback, processVariableMeasurements
            ]
            writer_medidas.writerow(new_row_medidas)

    # Informar ao usuário que o processo foi concluído
    print("Documento CSV de medidas criado com sucesso!")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Uso: python Criar Lista de Medidas.py arquivo_de_entrada_dados.xlsx nome_planilha diretorio_de_saida")
        sys.exit(1)

    arquivo_funcionalidades = sys.argv[1]
    nome_planilha = sys.argv[2]
    diretorio_saida = sys.argv[3]

    main(arquivo_funcionalidades, nome_planilha, diretorio_saida)
